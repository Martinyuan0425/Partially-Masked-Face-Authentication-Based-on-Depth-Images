import numpy as np
import openpyxl
from openpyxl import load_workbook
import cv2
import matplotlib.pyplot as plt
import math

def svd_compression(img, k):
    res_image = np.zeros_like(img)
    U, Sigma, VT = np.linalg.svd(img)
    sigma_250 = Sigma[:k]
    res_image = U[:,:k].dot(np.diag(Sigma[:k])).dot(VT[:k,:])

    return res_image, Sigma
    # return sigma_250

def excel(sigma, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NumPy_Arr"

    for x in sigma:
        myData = np.array([x])
        ws.append(myData.tolist())

    wb.save(filename=path)
    print("Saved", path)

def try_dis_3(X, state):
    th = [0.01, 0.03, 0.05, 0.05]
    sol = 0
    for j in range (3):
        dis = []
        for i in range(4):
            X_sum = (X[0][i]+X[1][i]+X[2][i])/3
            diff = abs((X[j][i]-X_sum)/X_sum)
            dis.append(diff)
        for i in range(len(dis)):
            if th[i]<=dis[i]:
                sol+=1
    
    # sol為設定是否開啟3禎偵測條件
    if state == False:
        sol = 0
        
    return sol

def try_dis(A):
    score = True
    B = [110000, 32000, 15000, 9000]
    th = [0.5, 0.5, 1, 1]
    dis = []
    for i in range(4):
        diff = (abs(A[i]-B[i])/A[i])
        dis.append(diff)
    
    for i in range(len(dis)):
        if th[i]<=dis[i]:
            score = False
            # print('label:', i, dis[i])
    return score

def load_sigma(path):
    S = []
    wb = load_workbook(path)
    sheet = wb['NumPy_Arr']
    for i in range(20):
        tag = 'A'+str(i+1)
        value = sheet[tag].value
        S.append(value)
    return S


    


